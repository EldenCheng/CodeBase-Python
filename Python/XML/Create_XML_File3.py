from xml.etree import ElementTree
import time
import datetime
import openpyxl
import shutil
from pathlib import Path
import random
from xml.dom.minidom import parse
import xml.dom.minidom

filenameNumber = 8
envelopeID = "7"
date = time.strftime("%Y-%m-%d", time.localtime())
time = time.strftime("%H:%M:%S", time.localtime())
PO_prefix = "STRESST"
PO_middle = "I"
PO_start_number = 50001

POs = 500
PO_items1 = 2
PO_items2 = 2
PO_items3 = 10
sea_air = "10"

#Open the xml file
tree = xml.dom.minidom.parse(r'.\Ori\PURORD_ECI_KERRY_StressTestSQL_001.xml')

#Create target Xml file
filepath = ".\Des\PURORD_ECI_KERRY_StressTest_%03d" % filenameNumber

while True:
    if Path(filepath + ".xml").is_file():
        filenameNumber = filenameNumber + 1
        filepath = ".\Des\PURORD_ECI_KERRY_StressTest_%03d" % filenameNumber
    else:
        filepath = ".\Des\PURORD_ECI_KERRY_StressTest_%03d" % filenameNumber
        break

#Create xlsx file to record the PO No and SKU No
shutil.copy(r'.\Ori\PO_Detail.xlsx', filepath + ".xlsx")
excelwb = openpyxl.load_workbook(filepath + ".xlsx")
excelst = excelwb.get_sheet_by_name("Sheet1")
SKUwb = openpyxl.load_workbook(r'.\Ori\SKUs.xlsx')
SKUst = SKUwb.get_sheet_by_name("Sheet1")

#Get the root node
root = tree.documentElement
Envelope = root.childNodes[1]
POM = root.getElementsByTagName("PurchaseOrderMessage")[0]
POM1 = POM.cloneNode(True)



#print(Envelope)
#print(POM)
#print(Envelope.getElementsByTagName("EnvelopeIdentification")[0].firstChild.nodeValue)
#print(Envelope.getElementsByTagName("Date"))
#print(Envelope.childNodes)


#Update Envelop itmes
Envelope.getElementsByTagName("EnvelopeIdentification")[0].firstChild.nodeValue = envelopeID
Envelope.getElementsByTagName("Date")[0].firstChild.nodeValue = date
Envelope.getElementsByTagName("Time")[0].firstChild.nodeValue = time

#PONO = PO_prefix + PO_middle + str(PO_start_number)
#POM.getElementsByTagName("PurchaseOrderNumber")[0].getElementsByTagName("Reference")[0].firstChild.nodeValue = PONO
#POM.getElementsByTagName("FOBDate")[0].firstChild.nodeValue = (datetime.datetime.now() +
#                                                               datetime.timedelta(days=7)).strftime('%Y-%m-%d')
#POM.getElementsByTagName("DateTimeInformation")[0].getElementsByTagName("Date")[0].firstChild.nodeValue = date
#excelst.cell(row=2, column=1).value = PONO

#print(root.childNodes)

#root.appendChild(root.childNodes[0].cloneNode(True))
#root.appendChild(POM.cloneNode(True))

#print(root.childNodes)

#print(POM.getElementsByTagName("PurchaseOrderNumber")[0].getElementsByTagName("Reference")[0].firstChild.nodeValue)


for i in range(0, POs):
    #POM = root.getElementsByTagName("PurchaseOrderMessage")[i]
    PONO = PO_prefix + PO_middle + str(PO_start_number + i)
    POM.getElementsByTagName("PurchaseOrderNumber")[0].getElementsByTagName("Reference")[0].firstChild.nodeValue = PONO
    POM.getElementsByTagName("FOBDate")[0].firstChild.nodeValue = (datetime.datetime.now() +
                                                                   datetime.timedelta(days=7)).strftime('%Y-%m-%d')
    POM.getElementsByTagName("DateTimeInformation")[0].getElementsByTagName("Date")[0].firstChild.nodeValue = date
    excelst.cell(row=(2 + i), column=1).value = PONO
    #Sea or Air
    if (i % 2) == 0:
        sea_air = "10"
        POM.getElementsByTagName("ModeOfTransport")[0].firstChild.nodeValue = sea_air
        excelst.cell(row=(2 + i), column=2).value = "Sea"
    else:
        sea_air = "40"
        POM.getElementsByTagName("ModeOfTransport")[0].firstChild.nodeValue = sea_air
        excelst.cell(row=(2 + i), column=2).value = "Air"

    if i < PO_items3:
        for j in range(PO_items1):
            print("first part i=%d, j=%d" % (i,j))
            PO_item = POM.getElementsByTagName("PurchaseOrderItem")[j]
            SKU_number = str(SKUst.cell(row=random.randint(2, 591), column=1).value)
            PO_item.getElementsByTagName("Identification")[0].firstChild.nodeValue = SKU_number
            excelst.cell(row=(2 + i), column=(3 + j)).value = SKU_number

            if j == 0:
                POM.appendChild(POM.childNodes[0].cloneNode(True))
                POM.appendChild(PO_item.cloneNode(True))
            elif j > 0 and (j != PO_items1 - 1):
                POM.appendChild(POM.childNodes[0].cloneNode(True))
                POM.appendChild(PO_item.cloneNode(True))
            else:
                POM.appendChild(POM.childNodes[0].cloneNode(True))
    else:
        for j in range(PO_items2):
            print("SECOND part i=%d, j=%d" % (i ,j))
            PO_item = POM.getElementsByTagName("PurchaseOrderItem")[j]
            SKU_number = str(SKUst.cell(row=random.randint(2, 591), column=1).value)
            PO_item.getElementsByTagName("Identification")[0].firstChild.nodeValue = SKU_number
            excelst.cell(row=(2 + i), column=(3 + j)).value = SKU_number

            if j != PO_items2 - 1:
                POM.appendChild(POM.childNodes[0].cloneNode(True))
                POM.appendChild(PO_item.cloneNode(True))
            else:
                POM.appendChild(POM.childNodes[0].cloneNode(True))

    if i == 0:
        root.appendChild(root.childNodes[0].cloneNode(True))
        POM = POM1.cloneNode(True)
        root.appendChild(POM.cloneNode(True))
        POM = root.getElementsByTagName("PurchaseOrderMessage")[i + 1]
    elif i > 0 and i != POs - 1:
        root.appendChild(root.childNodes[0].cloneNode(True))
        POM = POM1.cloneNode(True)
        root.appendChild(POM.cloneNode(True))
        POM = root.getElementsByTagName("PurchaseOrderMessage")[i + 1]
    else:
        root.appendChild(root.childNodes[0].cloneNode(True))

#Create a blank file
f = open(filepath + ".xml", "w")

#Save the xml content to the file
tree.writexml(f)

excelwb.save(filepath + ".xlsx")