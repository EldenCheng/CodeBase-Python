from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference


if __name__ == '__main__':

    """
    建立一个空白的Workbook,然后保存,就成为了一个空白的excel文件
    """
    wb = Workbook()
    wb.save("./test.xlsx")
    wb.close()

    """
    打开一个已存在的excel文件, 由于创建excel文件时,会同时创建一个sheet
    所以就算空白的Workbook也会有一个sheet
    """
    wb2 = load_workbook('test.xlsx')
    ws = wb2.active
    print(ws.title)
    wb2.close()


