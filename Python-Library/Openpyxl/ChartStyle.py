from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import series
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

wb = Workbook()
ws = wb.active

rows = [
    ("Sample",),
    (1,),
    (2,),
    (3,),
    (2,),
    (3,),
    (3,),
    (1,),
    (2,),
]

for r in rows:
    ws.append(r)


c = BarChart()
data = Reference(ws, min_col=1, min_row=1, max_row=8)
c.add_data(data, titles_from_data=True)
c.title = "Chart with patterns"

series = c.series[0]

#print(type(series))
print(series.graphicalProperties)
#print(c.path)

'''
1. DataPoint代表的就是Chart里的图形,这里是Bar Chart里的柱状图, idx是柱状图的序号,从0开始
2. graphicalProperties是柱状图的图形属性,有多种属性可以改,比如pattern就代表填充的pattern, solidFill就表示的就是填充的颜色等等
3. 填充颜色需要是ColorChoice的实例,颜色的代码可以是srgb,也可以是preset color等等, 
   preset color的值可以参考openpyxl包中的color.py里的定义
'''
pt = DataPoint(idx=0)
pt.graphicalProperties.solidFill = ColorChoice(prstClr="red")
series.dPt.append(pt)

pt = DataPoint(idx=1)
pt.graphicalProperties.solidFill = ColorChoice(prstClr="green")
series.dPt.append(pt)

'''
1. series本身都有graphicalProperties属性,也可以定义patternfill, solidFill之类的
2. pattFill就是整个series填充的模式,如果没有单独定义图柱的图形属性,那就默认与series的图形属性为准
'''
fill = PatternFillProperties(prst="pct5")
fill.foreground = ColorChoice(prstClr="red")
fill.background = ColorChoice(prstClr="blue")
series.graphicalProperties.pattFill = fill


ws.add_chart(c, "C1")

wb.save("pattern.xlsx")

'''
# set a pattern for the whole series
series = c.series[0]
fill = PatternFillProperties(prst="pct5")
fill.foreground = ColorChoice(prstClr="red")
fill.background = ColorChoice(prstClr="blue")
series.graphicalProperties.pattFill = fill

# set a pattern for a 6th data point (0-indexed)
pt = DataPoint(idx=5)
pt.graphicalProperties.pattFill = PatternFillProperties(prst="ltHorz")
series.dPt.append(pt)


ws.add_chart(c, "C1")
'''

#wb.save("pattern.xlsx")
wb.close()